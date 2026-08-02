"""Unit tests for detect_type_from_content covering filename + multi-sheet content detection."""
from __future__ import annotations

import os
import sys
from pathlib import Path

import pytest

sys.path.insert(0, "/app/backend")
from parsers import detect_type_from_content, detect_type_from_filename  # noqa: E402

FIX_CLEAN = Path("/app/backend/sample_fixtures/clean")
FIX_MESSY = Path("/app/backend/sample_fixtures/messy")


# ---------- Filename-only tests (fast path) ---------------------------------

@pytest.mark.parametrize("fname,expected", [
    ("pitch_deck.pdf", "pitch_deck"),
    ("mis_dashboard.xlsx", "mis"),
    ("financials.pdf", "financials"),
    ("projections_3yr.xlsx", "projections"),
    ("cap_table.csv", "cap_table"),
    ("Northwind_MIS_FY24.xlsx", "mis"),
    ("monthly_mis_dashboard.xlsx", "mis"),
    ("5yr_forecast.xlsx", "projections"),
    ("shareholding.csv", "cap_table"),
    ("Acme_AuditedFinancials.pdf", "financials"),
])
def test_filename_hint(fname, expected):
    got = detect_type_from_filename(fname)
    assert got == expected, f"{fname} -> {got}, expected {expected}"


# ---------- Real fixture files: filename + content ---------------------------

@pytest.mark.parametrize("fname,expected", [
    ("Northwind_PitchDeck_v3.pdf", "pitch_deck"),
    ("Northwind_MIS_FY24.xlsx", "mis"),
    ("Northwind_AuditedFinancials.pdf", "financials"),
    ("Northwind_Projections_3yr.xlsx", "projections"),
    ("Northwind_CapTable.csv", "cap_table"),
])
def test_clean_fixture_detection(fname, expected):
    path = FIX_CLEAN / fname
    assert path.exists(), path
    kind, conf = detect_type_from_content(fname, str(path))
    assert kind == expected, f"{fname}: got {kind}@{conf}"
    assert conf >= 0.75, f"{fname}: conf too low {conf}"


@pytest.mark.parametrize("fname,expected", [
    ("Acme_PitchDeck_Final.pdf", "pitch_deck"),
    ("Acme_MIS_Q4.xlsx", "mis"),
    ("Acme_Financials_2024.pdf", "financials"),
    ("Acme_Projections.xlsx", "projections"),
    ("Acme_CapTable.csv", "cap_table"),
])
def test_messy_fixture_detection(fname, expected):
    path = FIX_MESSY / fname
    assert path.exists(), path
    kind, conf = detect_type_from_content(fname, str(path))
    assert kind == expected, f"{fname}: got {kind}@{conf}"
    assert conf >= 0.75, f"{fname}: conf too low {conf}"


# ---------- Multi-sheet workbook with generic filename ---------------------

def _build_multi_sheet_xlsx(path: Path) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    s1 = wb.active
    s1.title = "MIS Dashboard"
    s1.append(["Month", "Revenue", "Burn", "CAC", "LTV"])
    s1.append(["Jan-24", 1000000, 200000, 500, 3000])
    s1.append(["Feb-24", 1100000, 210000, 490, 3100])
    s1.append(["Mar-24", 1150000, 220000, 480, 3200])
    s1.append(["Apr-24", 1200000, 230000, 470, 3300])
    s1.append(["May-24", 1250000, 240000, 460, 3400])
    s1.append(["Jun-24", 1300000, 250000, 450, 3500])

    s2 = wb.create_sheet("Projections FY26-27")
    s2.append(["Year", "Revenue Forecast", "Base Case", "Bull Case", "Bear Case"])
    s2.append(["FY26", 20000000, 20000000, 25000000, 15000000])
    s2.append(["FY27", 40000000, 40000000, 55000000, 25000000])

    s3 = wb.create_sheet("Notes")
    s3.append(["Note", "Detail"])
    s3.append(["Assumptions", "Growth 30% YoY"])

    wb.save(path)


def test_multi_sheet_generic_filename(tmp_path):
    p = tmp_path / "startup_report.xlsx"
    _build_multi_sheet_xlsx(p)
    kind, conf = detect_type_from_content("startup_report.xlsx", str(p))
    # Generic filename -> content wins. Must NOT be None/unknown.
    assert kind in {"mis", "projections"}, f"got {kind}@{conf}"
    assert conf >= 0.5, f"conf too low: {conf}"


def test_multi_sheet_mis_filename_wins(tmp_path):
    # Same workbook but named as MIS -> should confidently be mis
    p = tmp_path / "Company_MIS_Dashboard_FY26.xlsx"
    _build_multi_sheet_xlsx(p)
    kind, conf = detect_type_from_content(p.name, str(p))
    assert kind == "mis", f"got {kind}@{conf}"
    assert conf >= 0.9
